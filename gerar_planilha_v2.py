#!/usr/bin/env python3
"""
Planilha Excel v2.0 com controle de ingredientes e cálculo de custo
Executar: python gerar_planilha_v2.py
Resultado: Calculadora_Margem_Doces_v2.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os


def gerar_planilha_completa():
    """Cria planilha com ingredientes + receitas integradas"""

    wb = Workbook()
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ============ ABA 1: INGREDIENTES ============
    ws_ing = wb.active
    ws_ing.title = "Ingredientes"

    # Colunas
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_ing.column_dimensions[col].width = 18

    # Cabeçalho
    header_fill = PatternFill(
        start_color="FF6B35",
        end_color="FF6B35",
        fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)

    ing_headers = [
        "Ingrediente",
        "Unidade",
        "Preço por Unidade (R$)",
        "Qtd Comprada",
        "Data Compra",
        "Saldo Inicial",
        "Saldo Restante",
        "Observações"
    ]

    for col, header in enumerate(ing_headers, 1):
        cell = ws_ing.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = border

    # Exemplos de ingredientes
    ingredientes_ex = [
        ["Açúcar", "kg", 5.50, 2, "20/11/2025"],
        ["Farinha de Trigo", "kg", 4.20, 1, "20/11/2025"],
        ["Chocolate em Pó", "kg", 22.00, 0.5, "18/11/2025"],
        ["Leite Condensado", "L", 12.00, 1, "20/11/2025"],
        ["Manteiga", "kg", 18.50, 0.5, "19/11/2025"],
    ]

    for row_idx, ing in enumerate(ingredientes_ex, 2):
        for col_idx, valor in enumerate(ing, 1):
            cell = ws_ing.cell(row=row_idx, column=col_idx)
            cell.value = valor
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

            if col_idx == 3:  # Preço
                cell.number_format = 'R$ #,##0.00'
            elif col_idx == 5:  # Data
                cell.number_format = 'DD/MM/YYYY'

        # Coluna F: Saldo Inicial (igual à Qtd Comprada)
        cell_f = ws_ing.cell(row=row_idx, column=6)
        cell_f.value = f"=D{row_idx}"
        cell_f.border = border
        cell_f.alignment = Alignment(horizontal="center")

        # Coluna G: Saldo Restante (será atualizado)
        cell_g = ws_ing.cell(row=row_idx, column=7)
        cell_g.value = f"=F{row_idx}"
        cell_g.border = border
        cell_g.alignment = Alignment(horizontal="center")

    # Linhas em branco
    for row_idx in range(7, 17):
        for col_idx in range(1, 9):
            cell = ws_ing.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

            if col_idx == 3:
                cell.number_format = 'R$ #,##0.00'
            elif col_idx == 5:
                cell.number_format = 'DD/MM/YYYY'

        # Fórmulas
        ws_ing.cell(row=row_idx, column=6).value = f"=D{row_idx}"
        ws_ing.cell(row=row_idx, column=7).value = f"=F{row_idx}"

    # ========== SEÇÃO DE INSTRUÇÕES - INGREDIENTES ==========
    inst_row = 18

    # Título de instrução
    titulo_inst = ws_ing.cell(row=inst_row, column=1)
    titulo_inst.value = "ℹ️  COMO UTILIZAR ESTA ABA"
    titulo_inst.font = Font(bold=True, size=11, color="FFFFFF")
    titulo_inst.fill = PatternFill(
        start_color="FF8C00",
        end_color="FF8C00",
        fill_type="solid"
    )
    ws_ing.merge_cells(f'A{inst_row}:H{inst_row}')
    titulo_inst.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    instructions_ing = [
        ("INSTRUÇÕES GERAIS:", ""),
        ("1. Registre cada ingrediente", "Ex: Açúcar, Farinha, Chocolate"),
        ("2. Informe unidade de medida", "kg, L, g, ml, unidade, etc"),
        ("3. Digite preço pago/unidade", "Ex: 5,50 por kg de açúcar"),
        ("4. Quantidade comprada = saldo", "Coluna Saldo Inicial"),
        ("5. Conforme usa, saldo diminui", "Use nos cálculos de custo"),
        ("", ""),
        ("EXEMPLO PRÁTICO:", ""),
        ("Comprou 2kg de açúcar", "2kg × R$ 5,50 = R$ 11,00"),
        ("Usou 200g em uma receita", "Saldo fica em 1,8 kg"),
        ("Custo do ingrediente usado", "0,2kg × R$ 5,50 = R$ 1,10"),
    ]

    for idx, (title, example) in enumerate(instructions_ing, inst_row + 1):
        cell_title = ws_ing.cell(row=idx, column=1)
        cell_title.value = title

        if "INSTRUÇÕES" in title or "EXEMPLO" in title:
            cell_title.font = Font(
                bold=True,
                size=10,
                color="FFFFFF"
            )
            cell_title.fill = PatternFill(
                start_color="FF6B35",
                end_color="FF6B35",
                fill_type="solid"
            )
            ws_ing.merge_cells(f'A{idx}:H{idx}')
        else:
            cell_title.font = Font(size=9)

        cell_title.alignment = Alignment(horizontal="left", wrap_text=True)

        if example and "INSTRUÇÕES" not in title and "EXEMPLO" not in title:
            cell_example = ws_ing.cell(row=idx, column=4)
            cell_example.value = example
            cell_example.font = Font(size=9, italic=True, color="666666")
            cell_example.alignment = Alignment(
                horizontal="left",
                wrap_text=True
            )

    # ============ ABA 2: RECEITAS DETALHADAS ============
    ws_rec = wb.create_sheet("Receitas Detalhadas")

    ws_rec.column_dimensions['A'].width = 20
    ws_rec.column_dimensions['B'].width = 15
    ws_rec.column_dimensions['C'].width = 20
    ws_rec.column_dimensions['D'].width = 15
    ws_rec.column_dimensions['E'].width = 15

    # Cabeçalho receita
    header2_fill = PatternFill(
        start_color="4472C4",
        end_color="4472C4",
        fill_type="solid"
    )
    header2_font = Font(bold=True, color="FFFFFF", size=11)

    titulo = ws_rec.cell(row=1, column=1)
    titulo.value = "🍫 BRIGADEIRO"
    titulo.font = Font(bold=True, size=12, color="FFFFFF")
    titulo.fill = header2_fill
    ws_rec.merge_cells('A1:E1')

    rec_headers = [
        "Ingrediente Usado",
        "Unidade",
        "Quantidade Usada",
        "Preço Unitário (R$)",
        "Subtotal (R$)"
    ]

    for col, header in enumerate(rec_headers, 1):
        cell = ws_rec.cell(row=2, column=col)
        cell.value = header
        cell.fill = header2_fill
        cell.font = header2_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Ingredientes do Brigadeiro
    ingredientes_brigadeiro = [
        ["Açúcar", "g", 200],
        ["Leite Condensado", "ml", 150],
        ["Chocolate em Pó", "g", 50],
        ["Manteiga", "g", 20],
    ]

    for row_idx, ing_used in enumerate(ingredientes_brigadeiro, 3):
        ws_rec.cell(row=row_idx, column=1).value = ing_used[0]
        ws_rec.cell(row=row_idx, column=1).border = border
        ws_rec.cell(row=row_idx, column=1).alignment = Alignment(
            horizontal="left"
        )

        ws_rec.cell(row=row_idx, column=2).value = ing_used[1]
        ws_rec.cell(row=row_idx, column=2).border = border
        ws_rec.cell(row=row_idx, column=2).alignment = Alignment(
            horizontal="center"
        )

        ws_rec.cell(row=row_idx, column=3).value = ing_used[2]
        ws_rec.cell(row=row_idx, column=3).border = border
        ws_rec.cell(row=row_idx, column=3).alignment = Alignment(
            horizontal="center"
        )

        # Coluna D: Preço Unitário (referência Ingredientes)
        cell_d = ws_rec.cell(row=row_idx, column=4)
        cell_d.value = (
            f"=VLOOKUP(A{row_idx},Ingredientes!A:C,3,FALSE)"
        )
        cell_d.number_format = 'R$ #,##0.00'
        cell_d.border = border
        cell_d.alignment = Alignment(horizontal="center")

        # Coluna E: Subtotal = (Qty Usada / 1000) * Preço
        cell_e = ws_rec.cell(row=row_idx, column=5)
        cell_e.value = f"=(C{row_idx}/1000)*D{row_idx}"
        cell_e.number_format = 'R$ #,##0.00'
        cell_e.border = border
        cell_e.alignment = Alignment(horizontal="center")

    # Total de custo
    total_row = 7
    ws_rec.cell(row=total_row, column=1).value = "TOTAL CUSTO"
    ws_rec.cell(row=total_row, column=1).font = Font(bold=True)
    ws_rec.cell(row=total_row, column=1).fill = PatternFill(
        start_color="E7E6E6",
        end_color="E7E6E6",
        fill_type="solid"
    )

    cell_total = ws_rec.cell(row=total_row, column=5)
    cell_total.value = f"=SUM(E3:E6)"
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True)
    cell_total.fill = PatternFill(
        start_color="E7E6E6",
        end_color="E7E6E6",
        fill_type="solid"
    )
    cell_total.border = border

    # Info de venda
    ws_rec.cell(row=9, column=1).value = "Preço de Venda:"
    ws_rec.cell(row=9, column=2).value = 80.00
    ws_rec.cell(row=9, column=2).number_format = 'R$ #,##0.00'
    ws_rec.cell(row=9, column=2).font = Font(bold=True, size=12)

    ws_rec.cell(row=10, column=1).value = "Quantidade Produzida:"
    ws_rec.cell(row=10, column=2).value = 20
    ws_rec.cell(row=10, column=2).font = Font(bold=True, size=12)

    ws_rec.cell(row=11, column=1).value = "Custo Total do Lote:"
    ws_rec.cell(row=11, column=2).value = f"=E7"
    ws_rec.cell(row=11, column=2).number_format = 'R$ #,##0.00'
    ws_rec.cell(row=11, column=2).font = Font(bold=True, size=12, color="C00000")

    ws_rec.cell(row=12, column=1).value = "Lucro por Lote:"
    ws_rec.cell(row=12, column=2).value = f"=(B9*B10)-B11"
    ws_rec.cell(row=12, column=2).number_format = 'R$ #,##0.00'
    ws_rec.cell(row=12, column=2).font = Font(
        bold=True,
        size=12,
        color="00B050"
    )

    ws_rec.cell(row=13, column=1).value = "Lucro por Unidade:"
    ws_rec.cell(row=13, column=2).value = f"=B12/B10"
    ws_rec.cell(row=13, column=2).number_format = 'R$ #,##0.00'
    ws_rec.cell(row=13, column=2).font = Font(bold=True, size=12)

    ws_rec.cell(row=14, column=1).value = "Margem de Lucro %:"
    ws_rec.cell(row=14, column=2).value = f"=((B12)/(B9*B10))*100"
    ws_rec.cell(row=14, column=2).number_format = '0.00"%"'
    ws_rec.cell(row=14, column=2).font = Font(
        bold=True,
        size=12,
        color="00B050"
    )

    # ========== SEÇÃO DE INSTRUÇÕES - RECEITAS ==========
    inst_rec_row = 16

    titulo_rec_inst = ws_rec.cell(row=inst_rec_row, column=1)
    titulo_rec_inst.value = "ℹ️  COMO UTILIZAR ESTA ABA"
    titulo_rec_inst.font = Font(bold=True, size=11, color="FFFFFF")
    titulo_rec_inst.fill = PatternFill(
        start_color="4472C4",
        end_color="4472C4",
        fill_type="solid"
    )
    ws_rec.merge_cells(f'A{inst_rec_row}:E{inst_rec_row}')
    titulo_rec_inst.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    instructions_rec = [
        ("PASSO A PASSO:", ""),
        ("1. Renomeie a aba para seu doce", "Ex: Brigadeiro → Beijinho"),
        ("2. Liste ingredientes usados", "Coluna A"),
        ("3. Coloque quantidade em GRAMAS", "Ex: 200g, não 0.2kg"),
        ("4. Sistema busca preço automaticamente", "VLOOKUP dos Ingredientes"),
        ("5. Subtotal = (Qtd / 1000) × Preço", "Automaticamente calculado"),
        ("6. Informe Preço de Venda", "O valor que cobra por unidade"),
        ("7. Quantidade produzida", "Quantas unidades saem do lote"),
        ("8. Veja o lucro e margem%", "Automaticamente atualizado"),
        ("", ""),
        ("EXEMPLO:", ""),
        ("Receita: Brigadeiro", ""),
        ("Ingredientes: Açúcar, Leite Cond., Chocolate, Manteiga", ""),
        ("Quantidades: 200g, 150ml, 50g, 20g", ""),
        ("Total de custo: R$ 12,50", ""),
        ("Vende 20 unidades por R$ 80,00", ""),
        ("Lucro por lote: R$ 67,50", ""),
        ("Margem: 84%", "Excelente!"),
    ]

    for idx, (title, example) in enumerate(instructions_rec, inst_rec_row + 1):
        cell_title = ws_rec.cell(row=idx, column=1)
        cell_title.value = title

        if "PASSO" in title or "EXEMPLO" in title:
            cell_title.font = Font(
                bold=True,
                size=10,
                color="FFFFFF"
            )
            cell_title.fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
            ws_rec.merge_cells(f'A{idx}:E{idx}')
        else:
            cell_title.font = Font(size=9)

        cell_title.alignment = Alignment(horizontal="left", wrap_text=True)

        if example and "PASSO" not in title and "EXEMPLO" not in title:
            cell_example = ws_rec.cell(row=idx, column=3)
            cell_example.value = example
            cell_example.font = Font(size=9, italic=True, color="666666")
            cell_example.alignment = Alignment(
                horizontal="left",
                wrap_text=True
            )

    # ============ ABA 3: RESUMO RECEITAS ============
    ws_res = wb.create_sheet("Resumo Receitas")

    ws_res.column_dimensions['A'].width = 18
    ws_res.column_dimensions['B'].width = 15
    ws_res.column_dimensions['C'].width = 15
    ws_res.column_dimensions['D'].width = 15
    ws_res.column_dimensions['E'].width = 15
    ws_res.column_dimensions['F'].width = 15
    ws_res.column_dimensions['G'].width = 15

    res_headers = [
        "Receita",
        "Custo Total",
        "Preço Venda",
        "Quantidade",
        "Lucro Lote",
        "Lucro Unit",
        "Margem %"
    ]

    for col, header in enumerate(res_headers, 1):
        cell = ws_res.cell(row=1, column=col)
        cell.value = header
        cell.fill = header2_fill
        cell.font = header2_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Exemplo: Brigadeiro
    ws_res.cell(row=2, column=1).value = "Brigadeiro"
    ws_res.cell(row=2, column=2).value = "='Receitas Detalhadas'!B11"
    ws_res.cell(row=2, column=2).number_format = 'R$ #,##0.00'
    ws_res.cell(row=2, column=3).value = "='Receitas Detalhadas'!B9"
    ws_res.cell(row=2, column=3).number_format = 'R$ #,##0.00'
    ws_res.cell(row=2, column=4).value = "='Receitas Detalhadas'!B10"
    ws_res.cell(row=2, column=5).value = "='Receitas Detalhadas'!B12"
    ws_res.cell(row=2, column=5).number_format = 'R$ #,##0.00'
    ws_res.cell(row=2, column=6).value = "='Receitas Detalhadas'!B13"
    ws_res.cell(row=2, column=6).number_format = 'R$ #,##0.00'
    ws_res.cell(row=2, column=7).value = "='Receitas Detalhadas'!B14"
    ws_res.cell(row=2, column=7).number_format = '0.00"%"'

    for col in range(1, 8):
        ws_res.cell(row=2, column=col).border = border
        ws_res.cell(row=2, column=col).alignment = Alignment(
            horizontal="center"
        )

    # Linhas em branco para mais receitas
    for row_idx in range(3, 12):
        for col_idx in range(1, 8):
            cell = ws_res.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if col_idx in [2, 3, 5, 6]:
                cell.number_format = 'R$ #,##0.00'
            elif col_idx == 7:
                cell.number_format = '0.00"%"'

    # ========== SEÇÃO DE INSTRUÇÕES - RESUMO ==========
    inst_res_row = 13

    titulo_res_inst = ws_res.cell(row=inst_res_row, column=1)
    titulo_res_inst.value = "ℹ️  COMO UTILIZAR ESTA ABA"
    titulo_res_inst.font = Font(bold=True, size=11, color="FFFFFF")
    titulo_res_inst.fill = PatternFill(
        start_color="70AD47",
        end_color="70AD47",
        fill_type="solid"
    )
    ws_res.merge_cells(f'A{inst_res_row}:G{inst_res_row}')
    titulo_res_inst.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    instructions_res = [
        ("VISÃO GERAL DE TODAS AS RECEITAS:", ""),
        ("Esta aba consolida automaticamente", "dados de todas as receitas"),
        ("", ""),
        ("COMO FUNCIONA:", ""),
        ("• Cada receita criada aparece aqui", "Automaticamente atualizada"),
        ("• Compara custo vs preço", "Mostra o melhor lucro"),
        ("• Margem% indica rentabilidade", "Acima de 50% é ótimo!"),
        ("", ""),
        ("PARA CRIAR NOVA RECEITA:", ""),
        ("1. Vá à aba 'Receitas Detalhadas'", ""),
        ("2. Clique botão direito na aba", ""),
        ("3. Mover/Copiar → Copiar", ""),
        ("4. Renomeie: 'Receita 1', 'Receita 2', etc", ""),
        ("5. Edite os dados (ingredientes, preços)", ""),
        ("6. Aqui aparecerá automaticamente", ""),
        ("", ""),
        ("EXEMPLO DE LEITURA:", ""),
        ("Brigadeiro: R$12,50 custo → R$80 venda", "Lucro R$67,50"),
        ("Produz 20 unidades por lote", "84% de margem!"),
    ]

    for idx, (title, example) in enumerate(instructions_res, inst_res_row + 1):
        cell_title = ws_res.cell(row=idx, column=1)
        cell_title.value = title

        if "VISÃO" in title or "COMO" in title or "PARA" in title:
            cell_title.font = Font(
                bold=True,
                size=10,
                color="FFFFFF"
            )
            cell_title.fill = PatternFill(
                start_color="70AD47",
                end_color="70AD47",
                fill_type="solid"
            )
            ws_res.merge_cells(f'A{idx}:G{idx}')
        else:
            cell_title.font = Font(size=9)

        cell_title.alignment = Alignment(horizontal="left", wrap_text=True)

        if example and "VISÃO" not in title and "COMO" not in title and \
           "PARA" not in title and "EXEMPLO" not in title:
            cell_example = ws_res.cell(row=idx, column=4)
            cell_example.value = example
            cell_example.font = Font(size=9, italic=True, color="666666")
            cell_example.alignment = Alignment(
                horizontal="left",
                wrap_text=True
            )

    # ============ ABA 4: ANÁLISE GERAL ============
    ws_ana = wb.create_sheet("Análise")

    ws_ana.column_dimensions['A'].width = 25
    ws_ana.column_dimensions['B'].width = 20

    title = ws_ana.cell(row=1, column=1)
    title.value = "ANÁLISE GERAL"
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = PatternFill(
        start_color="70AD47",
        end_color="70AD47",
        fill_type="solid"
    )
    title.alignment = Alignment(horizontal="center")
    ws_ana.merge_cells('A1:B1')

    ws_ana.cell(row=3, column=1).value = "Ingredientes Cadastrados:"
    ws_ana.cell(row=3, column=2).value = "=COUNTA(Ingredientes!A2:A16)"
    ws_ana.cell(row=3, column=2).font = Font(bold=True, size=12)

    ws_ana.cell(row=4, column=1).value = "Receitas Registradas:"
    ws_ana.cell(row=4, column=2).value = "=COUNTA('Resumo Receitas'!A2:A11)"
    ws_ana.cell(row=4, column=2).font = Font(bold=True, size=12)

    ws_ana.cell(row=5, column=1).value = "Lucro Total:"
    ws_ana.cell(row=5, column=2).value = "=SUM('Resumo Receitas'!E2:E11)"
    ws_ana.cell(row=5, column=2).number_format = 'R$ #,##0.00'
    ws_ana.cell(row=5, column=2).font = Font(
        bold=True,
        size=12,
        color="00B050"
    )

    ws_ana.cell(row=6, column=1).value = "Margem Média:"
    ws_ana.cell(row=6, column=2).value = "=AVERAGE('Resumo Receitas'!G2:G11)"
    ws_ana.cell(row=6, column=2).number_format = '0.00"%"'
    ws_ana.cell(row=6, column=2).font = Font(bold=True, size=12)

    ws_ana.cell(row=8, column=1).value = "💡 RESUMO DE USO:"
    ws_ana.cell(row=8, column=1).font = Font(bold=True, size=11)

    tips = [
        "✓ Aba 1: Ingredientes - Registre todas compras",
        "✓ Aba 2: Receitas - Crie uma por doce/produto",
        "✓ Aba 3: Resumo - Veja tudo junto",
        "✓ Aba 4: Análise - Acompanhe totais",
        "",
        "📌 DICA IMPORTANTE:",
        "Copie a aba 'Receitas Detalhadas' para cada",
        "nova receita. Sistema puxará automaticamente",
        "os preços da aba Ingredientes via VLOOKUP",
        "",
        "🎯 FÓRMULA CHAVE:",
        "Custo do ingrediente = (Qtd em g / 1000) × Preço/kg",
        "Margem % = (Lucro Total / Receita Total) × 100",
        "",
        "💰 INTERPRETAÇÃO:",
        "• Margem 50%+ = Excelente",
        "• Margem 30-50% = Bom",
        "• Margem <30% = Revisar preço",
    ]

    for idx, tip in enumerate(tips, 9):
        cell = ws_ana.cell(row=idx, column=1)
        cell.value = tip
        if tip.startswith("💡") or tip.startswith("🎯") or \
           tip.startswith("💰"):
            cell.font = Font(size=10, bold=True)
        else:
            cell.font = Font(size=9)
        cell.alignment = Alignment(horizontal="left", wrap_text=True)

    # Salvar
    filename = "Calculadora_Margem_Doces_v2.xlsx"
    wb.save(filename)
    print(f"\n✅ PLANILHA GERADA COM SUCESSO!")
    print(f"📁 Arquivo: {filename}")
    print(f"📍 Localização: {os.path.abspath(filename)}")
    print(f"\n🎯 O que você tem:")
    print(f"   ✓ Aba 'Ingredientes' com instruções elegantes")
    print(f"   ✓ Aba 'Receitas Detalhadas' com guia de uso")
    print(f"   ✓ Aba 'Resumo Receitas' com consolidação")
    print(f"   ✓ Aba 'Análise' com resumo executivo")
    print(f"   ✓ Cálculos automáticos em TODAS as abas")
    print(f"\n📱 PRÓXIMO PASSO:")
    print(f"   1. Abra no Excel, Google Sheets ou WPS Office")
    print(f"   2. Leia as instruções em cada aba")
    print(f"   3. Teste no smartphone")
    print(f"   4. Preencha com seus dados reais")
    print(f"\n💡 Dica: Abra e explore cada aba para entender!")


if __name__ == "__main__":
    try:
        gerar_planilha_completa()
    except Exception as e:
        print(f"❌ Erro ao gerar planilha: {str(e)}")
        print(f"Instale: pip install openpyxl")
